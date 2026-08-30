"""Pure, request-scoped DOORS script generation endpoints."""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from awcenter.api_errors import error_response
from awcenter.file_security import OOXML_WORKBOOK_POLICY, validate_request_upload

from integrations.doors.escape import dxl_quote
from .serializers import ScriptGenerationSerializer

MAX_WORKBOOK_COLUMNS = 1_000
MAX_WORKBOOK_ROWS = 10_000
MAX_CELL_CHARACTERS = 32_767
MAX_SCRIPT_BYTES = 5 * 1024 * 1024
MAX_SCRIPT_SOURCE_BYTES = 2 * 1024 * 1024


class ScriptGenerationError(ValueError):
    """Represent a safe, user-actionable workbook conversion failure."""

    def __init__(self, detail, code):
        super().__init__(detail)
        self.detail = detail
        self.code = code


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_script(request):
    """Create bounded, escaped DXL from the first sheet of an OOXML workbook."""

    excel_file = validate_request_upload(request, "file", OOXML_WORKBOOK_POLICY)
    serializer = ScriptGenerationSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    try:
        result = generate_dxl_script(excel_file, serializer.validated_data["json"])
    except ScriptGenerationError as error:
        return error_response(error.detail, error.code, response_status=400)
    return Response(result)


def generate_dxl_script(excel_file, mappings):
    """Read one bounded worksheet and return a deterministic JSON-safe result."""

    ordered_mappings = sorted(mappings, key=lambda item: not item["search"])
    rows = read_mapped_rows(excel_file, ordered_mappings)
    arrays = build_arrays(ordered_mappings, rows)
    assignments = build_assignments(ordered_mappings)
    search_attribute = dxl_quote(ordered_mappings[0]["doors"])
    script = f'''#include <addins/user/yck.dxl>

Module refModule = current
if (null refModule) {{
    print "Module not found\\n"
    halt
}}

{arrays}

Object o
int i = 0
for (i = 0; i < {len(rows)}; i++) {{
    o = FindObjectByAttribute(refModule, {search_attribute}, awc_arr_1[i])
    if (null o) {{
        print "Object not found: " awc_arr_1[i] "\\n"
    }}
    else {{
{assignments}
    }}
}}
'''
    if len(script.encode("utf-8")) > MAX_SCRIPT_BYTES:
        raise ScriptGenerationError(
            "The generated script exceeds the safety limit.",
            "DOORS_SCRIPT_SIZE_LIMIT",
        )
    return {
        "script": script,
        "row_count": len(rows),
        "mapping_count": len(ordered_mappings),
    }


def read_mapped_rows(excel_file, mappings):
    """Return bounded scalar rows without executing workbook formulas or macros."""

    from openpyxl import load_workbook

    try:
        excel_file.seek(0)
        workbook = load_workbook(
            excel_file,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as error:
        raise ScriptGenerationError(
            "The workbook could not be read.",
            "DOORS_WORKBOOK_INVALID",
        ) from error

    try:
        worksheet = workbook.worksheets[0]
        validate_sheet_dimensions(worksheet)
        headers = read_headers(worksheet)
        positions = resolve_mapping_positions(headers, mappings)
        rows = []
        source_bytes = 0
        maximum_column = max(positions) + 1
        for source_row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
            max_col=maximum_column,
            values_only=True,
        ):
            row = [normalize_cell(source_row[position]) for position in positions]
            if any(row):
                source_bytes += sum(len(value.encode("utf-8")) for value in row)
                if source_bytes > MAX_SCRIPT_SOURCE_BYTES:
                    raise ScriptGenerationError(
                        "The generated script exceeds the safety limit.",
                        "DOORS_SCRIPT_SIZE_LIMIT",
                    )
                rows.append(row)
        if not rows:
            raise ScriptGenerationError(
                "The workbook has no mapped data rows.",
                "DOORS_WORKBOOK_EMPTY",
            )
        return rows
    except ScriptGenerationError:
        raise
    except Exception as error:
        raise ScriptGenerationError(
            "The workbook could not be converted.",
            "DOORS_WORKBOOK_INVALID",
        ) from error
    finally:
        workbook.close()
        excel_file.seek(0)


def validate_sheet_dimensions(worksheet):
    """Reject sheets whose declared dimensions exceed conversion limits."""

    if worksheet.max_column < 1 or worksheet.max_row < 1:
        raise ScriptGenerationError("The workbook is empty.", "DOORS_WORKBOOK_EMPTY")
    if (
        worksheet.max_column > MAX_WORKBOOK_COLUMNS
        or worksheet.max_row - 1 > MAX_WORKBOOK_ROWS
    ):
        raise ScriptGenerationError(
            "The workbook exceeds the script generation limits.",
            "DOORS_WORKBOOK_SIZE_LIMIT",
        )


def read_headers(worksheet):
    """Read and normalize the first row as unique column names."""

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    present_headers = [header for header in headers if header]
    if not present_headers:
        raise ScriptGenerationError("The workbook has no headers.", "DOORS_WORKBOOK_EMPTY")
    normalized = [header.casefold() for header in present_headers]
    if len(normalized) != len(set(normalized)):
        raise ScriptGenerationError(
            "Workbook column names must be unique.",
            "DOORS_WORKBOOK_HEADERS_INVALID",
        )
    return headers


def resolve_mapping_positions(headers, mappings):
    """Resolve case-insensitive mappings to stable worksheet column indexes."""

    lookup = {header.casefold(): index for index, header in enumerate(headers) if header}
    try:
        return [lookup[item["excel"].casefold()] for item in mappings]
    except KeyError as error:
        raise ScriptGenerationError(
            "The workbook headers do not match the submitted mappings.",
            "DOORS_SCRIPT_MAPPING_INVALID",
        ) from error


def normalize_cell(value):
    """Convert a workbook scalar to one bounded DXL string value."""

    text = "" if value is None else str(value).strip()
    if len(text) > MAX_CELL_CHARACTERS:
        raise ScriptGenerationError(
            "A workbook cell exceeds the script generation limit.",
            "DOORS_WORKBOOK_CELL_LIMIT",
        )
    return text


def build_arrays(mappings, rows):
    """Build escaped DXL arrays for each ordered mapping."""

    arrays = []
    for index in range(len(mappings)):
        values = ",".join(dxl_quote(row[index]) for row in rows)
        arrays.append(f"string awc_arr_{index + 1}[] = {{{values}}}")
    return "\n".join(arrays)


def build_assignments(mappings):
    """Build escaped attribute writes while keeping the search mapping read-only."""

    if len(mappings) == 1:
        return "        // No attribute updates were selected."
    return "\n".join(
        f"        SetObjectAttribute(o, {dxl_quote(item['doors'])}, awc_arr_{index}[i])"
        for index, item in enumerate(mappings[1:], start=2)
    )
