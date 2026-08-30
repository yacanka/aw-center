import json
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook, load_workbook
from rest_framework.test import APITestCase


class ExcelComparisonSecurityTests(APITestCase):
    url = "/api/tools/excel/compare/"

    def setUp(self):
        self.user = get_user_model().objects.create_user("excel-user", password="runtime-test-only")
        self.client.force_authenticate(self.user)

    def test_duplicate_keys_return_bounded_error_without_row_contents(self):
        first = workbook([["id", "name"], ["1", "private-row"], ["1", "private-row"]])
        second = workbook([["id", "name"], ["1", "updated"]])

        response = self.compare(first, second)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.data["code"], "EXCEL_COMPARE_DUPLICATE_KEYS")
        self.assertNotIn("private-row", str(response.data))

    def test_untrusted_formula_is_neutralized_in_comparison_workbook(self):
        first = workbook([["id", "name"], ["1", "Existing"]])
        # A leading plus remains text in the uploaded workbook but is still a
        # spreadsheet formula trigger when written back to Excel.
        second = workbook([["id", "name"], ["1", "Existing"], ["2", "+1+1"]])

        response = self.compare(first, second)

        self.assertEqual(response.status_code, 200)
        added_sheet = load_workbook(BytesIO(response.content), data_only=False)["Added"]
        self.assertNotEqual(added_sheet["B2"].data_type, "f")
        self.assertEqual(added_sheet["B2"].value, "'+1+1")

    def compare(self, first, second):
        return self.client.post(
            self.url,
            {
                "first": first,
                "second": second,
                "json": json.dumps({"keyColumns": ["id"]}),
            },
            format="multipart",
        )


def workbook(rows):
    output = BytesIO()
    document = Workbook()
    sheet = document.active
    for row in rows:
        sheet.append(row)
    document.save(output)
    return SimpleUploadedFile(
        "comparison.xlsx",
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
