"""Styled Excel export generation for project CompDoc models."""

from io import BytesIO

import pandas as pd
from django.http import HttpResponse
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView

from .compdoc_permissions import StrictDjangoModelPermissions

SHEET_NAME = "Compliance Documents"
MAX_COLUMN_WIDTH = 50
LIST_COLUMNS = {"Signature Panel", "Requirements", "Status Flow"}
WRAP_COLUMNS = LIST_COLUMNS | {"Tech Doc No", "Tech Doc Issue", "Delivered Tech Doc Issue"}
SECONDARY_FIELDS = (
    ("tech_doc_no", "tech_doc_no_2"),
    ("tech_doc_issue", "tech_doc_issue_2"),
    ("delivered_tech_doc_issue", "delivered_tech_doc_issue_2"),
)


def excel_creator_factory(model, serializer_class, view_permission_classes):
    """Return a project-specific CompDoc Excel export API view."""

    class ExcelCreator(APIView):
        """Download the current project CompDocs as a styled workbook."""

        queryset = model.objects.none()
        permission_classes = [*view_permission_classes, StrictDjangoModelPermissions]

        def get(self, request):
            """Build and return one in-memory OOXML workbook."""

            return build_excel_response(model, serializer_class)

    return ExcelCreator


def build_excel_response(model, serializer_class):
    """Serialize model rows and return a downloadable workbook response."""

    serialized_rows = serializer_class(model.objects.all(), many=True).data
    dataframe = prepare_export_dataframe(pd.DataFrame(serialized_rows))
    buffer = write_workbook(dataframe)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="Compliance Documents.xlsx"'
    return response


def prepare_export_dataframe(dataframe):
    """Return a human-readable export dataframe without internal fields."""

    for primary, secondary in SECONDARY_FIELDS:
        merge_secondary_column(dataframe, primary, secondary)
    add_status_columns(dataframe)
    dataframe.drop(columns=["id", "path", "created_time"], errors="ignore", inplace=True)
    dataframe.columns = [str(column).replace("_", " ").title() for column in dataframe.columns]
    normalize_list_columns(dataframe)
    return dataframe


def merge_secondary_column(dataframe, primary, secondary):
    """Merge an optional secondary document value into its primary column."""

    if secondary not in dataframe.columns:
        return
    dataframe[primary] = dataframe.apply(
        lambda row: join_present_values(row.get(primary), row.get(secondary)), axis=1
    )
    dataframe.drop(columns=[secondary], inplace=True)


def join_present_values(primary, secondary):
    """Join non-null scalar values using the established newline format."""

    values = [value for value in (primary, secondary) if value is not None and not pd.isna(value)]
    return "\n".join(str(value) for value in values)


def add_status_columns(dataframe):
    """Derive target, delivery, and current status from status history."""

    if "status_flow" not in dataframe.columns:
        return
    dataframe["ubm_target_date"] = dataframe["status_flow"].apply(
        lambda flow: status_date(flow, "to_be_issued")
    )
    dataframe["ubm_delivery_date"] = dataframe["status_flow"].apply(
        lambda flow: status_date(flow, "authority_review")
    )
    dataframe["status"] = dataframe["status_flow"].apply(current_status)


def status_date(flow, status):
    """Return the first matching status date from a safe event list."""

    if not isinstance(flow, list):
        return None
    return next((event.get("date") for event in flow if event.get("status") == status), None)


def current_status(flow):
    """Return the last status identifier from a safe event list."""

    if not isinstance(flow, list) or not flow:
        return None
    return flow[-1].get("status")


def normalize_list_columns(dataframe):
    """Render JSON list columns as newline-separated workbook cells."""

    for column in LIST_COLUMNS.intersection(dataframe.columns):
        dataframe[column] = dataframe[column].apply(format_list_value)


def format_list_value(value):
    """Return a stable workbook representation for a JSON-list value."""

    if isinstance(value, (list, tuple)):
        return "\n".join(map(str, value))
    return "" if value is None or pd.isna(value) else str(value)


def write_workbook(dataframe):
    """Write and style one in-memory workbook buffer."""

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        style_worksheet(writer.sheets[SHEET_NAME], dataframe.columns)
    buffer.seek(0)
    return buffer


def style_worksheet(worksheet, columns):
    """Apply bounded widths, stripes, wrapping, and header styling."""

    set_column_widths(worksheet)
    add_row_striping(worksheet)
    add_wrapping(worksheet, columns)
    style_header(worksheet)


def set_column_widths(worksheet):
    """Set readable widths capped against oversized cell values."""

    for column_index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(column_index)
        lengths = [len(str(cell.value or "")) for cell in worksheet[letter]]
        worksheet.column_dimensions[letter].width = min(max(lengths, default=0) + 2, MAX_COLUMN_WIDTH)


def add_row_striping(worksheet):
    """Apply alternating fill across the complete data column range."""

    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    last_column = get_column_letter(worksheet.max_column)
    fill = PatternFill(start_color="87CEB3", end_color="87CEB3", fill_type="solid")
    worksheet.conditional_formatting.add(
        f"A2:{last_column}{worksheet.max_row}",
        FormulaRule(formula=["MOD(ROW(),2)=0"], fill=fill),
    )


def add_wrapping(worksheet, columns):
    """Wrap only established multiline export columns."""

    for column_index, column in enumerate(columns, start=1):
        if column not in WRAP_COLUMNS:
            continue
        for cell in worksheet[get_column_letter(column_index)]:
            cell.alignment = Alignment(wrap_text=True)


def style_header(worksheet):
    """Apply the established dark export header styling."""

    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="5BCEA8")
        cell.fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")
