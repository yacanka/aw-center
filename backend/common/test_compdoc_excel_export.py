from io import BytesIO
from zipfile import ZipFile

import pandas as pd
from django.apps import apps
from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import load_workbook

from .compdoc_excel_export import (
    build_excel_response,
    prepare_export_dataframe,
    write_workbook,
)
from .compdoc_import import get_missing_required_fields, map_headers
from .compdoc_import_values import get_mappable_import_fields, normalize_import_row
from projects.ozgur.models import CompDoc
from projects.ozgur.serializers import CompDocSerializer
from projects.registry import PROJECT_DEFINITIONS


class CompDocExcelExportTests(SimpleTestCase):
    """Verify safe export normalization and complete worksheet styling."""

    def test_export_merges_secondary_values_and_derives_status(self):
        """Internal fields are removed while workflow values remain readable."""

        dataframe = pd.DataFrame(
            [
                {
                    "id": "private-id",
                    "path": "/private/path",
                    "created_time": "2026-01-01",
                    "tech_doc_no": "TD-1",
                    "tech_doc_no_2": "TD-2",
                    "requirements": ["REQ-1", "REQ-2"],
                    "status_flow": [
                        {"status": "to_be_issued", "date": "01.01.2026"},
                        {"status": "authority_approved", "date": "02.01.2026"},
                    ],
                }
            ]
        )

        result = prepare_export_dataframe(dataframe, CompDoc)

        self.assertNotIn("Id", result.columns)
        self.assertNotIn("Path", result.columns)
        self.assertEqual(result.loc[0, "Tech Doc No"], "TD-1\nTD-2")
        self.assertEqual(result.loc[0, "Requirements"], "REQ-1\nREQ-2")
        self.assertEqual(result.loc[0, "Status"], "authority_approved")
        self.assertTrue(result.loc[0, "Status Flow"].startswith('{"status":"to_be_issued"'))

    def test_workbook_is_readable_navigable_and_editable(self):
        """The canonical data sheet includes operator guidance and validation."""

        dataframe = pd.DataFrame(
            [
                {
                    "Name": "Manual",
                    "Panel": "Flight",
                    "Status": "to_be_issued",
                    "Cat": "1",
                    "Moc": "0",
                }
            ]
        )
        workbook = load_workbook(BytesIO(write_workbook(dataframe).getvalue()))
        worksheet = workbook["Compliance Documents"]

        self.assertEqual(workbook.sheetnames, ["Compliance Documents", "Dashboard"])
        self.assertEqual(workbook.active.title, "Dashboard")
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertIsNone(worksheet.auto_filter.ref)
        self.assertIn("ComplianceDocumentsTable", worksheet.tables)
        self.assertEqual(len(worksheet.data_validations.dataValidation), 3)
        dashboard = workbook["Dashboard"]
        self.assertTrue(dashboard["A6"].value.startswith("=COUNTA("))
        self.assertEqual(dashboard["M6"].number_format, "0%")
        self.assertEqual(len(dashboard._charts), 2)
        for validation in worksheet.data_validations.dataValidation:
            self.assertTrue(validation.formula1.startswith('"'))
            self.assertNotIn("!", validation.formula1)

    def test_table_owns_the_only_filter_definition(self):
        """Avoid Excel repairs caused by overlapping worksheet and table filters."""

        content = write_workbook(pd.DataFrame([{"Name": "Manual", "Status": "unknown"}])).getvalue()

        with ZipFile(BytesIO(content)) as package:
            worksheet_xml = package.read("xl/worksheets/sheet1.xml")
            table_xml = package.read("xl/tables/table1.xml")

        self.assertNotIn(b"<autoFilter", worksheet_xml)
        self.assertEqual(table_xml.count(b"<autoFilter"), 1)

    def test_every_project_exports_one_fully_importable_sheet(self):
        """Even empty project exports expose only headers accepted by their importer."""

        for project_slug, definition in PROJECT_DEFINITIONS.items():
            if "compdocs" not in definition.capabilities:
                continue
            with self.subTest(project=project_slug):
                model = apps.get_model(project_slug, "CompDoc")
                exported = prepare_export_dataframe(pd.DataFrame(), model)
                content = write_workbook(exported).getvalue()
                workbook = load_workbook(BytesIO(content))
                loaded = pd.read_excel(BytesIO(content))
                importable = get_mappable_import_fields(model)
                mapping = map_headers(loaded.columns, importable)

                self.assertEqual(workbook.sheetnames, ["Compliance Documents", "Dashboard"])
                self.assertEqual(set(mapping), set(loaded.columns))
                self.assertEqual(get_missing_required_fields(mapping.values(), importable), [])

    def test_export_round_trips_through_the_import_contract(self):
        """Exported headers and multiline values normalize without manual cleanup."""

        source = pd.DataFrame(
            [
                {
                    "name": "Flight Manual",
                    "panel": "Flight",
                    "responsible": "Reviewer",
                    "ata": "21-00",
                    "status": "authority_approved",
                    "cat": "1",
                    "moc": "0",
                    "cover_page_no": "CP-1",
                    "cover_page_issue": "2",
                    "tech_doc_no": "TD-1",
                    "tech_doc_issue": "3",
                    "delivered_tech_doc_issue": None,
                    "requirements": ["REQ-1", "REQ-2"],
                    "status_flow": [
                        {"status": "to_be_issued", "date": "01.01.2026"},
                        {"status": "authority_approved", "date": "02.01.2026"},
                    ],
                }
            ]
        )
        exported = prepare_export_dataframe(source, CompDoc)
        loaded = pd.read_excel(BytesIO(write_workbook(exported).getvalue()))
        mapping = map_headers(loaded.columns, get_mappable_import_fields(CompDoc))
        normalized = normalize_import_row(loaded.rename(columns=mapping).iloc[0].to_dict(), CompDoc)

        self.assertEqual(normalized["cover_page_no"], "CP-1")
        self.assertEqual(normalized["requirements"], ["REQ-1", "REQ-2"])
        self.assertIsNone(normalized["delivered_tech_doc_issue"])
        self.assertEqual(normalized["status_flow"][1]["status"], "authority_approved")

    def test_edited_export_columns_reconcile_without_losing_history(self):
        """Status and milestone edits append/update exported history deterministically."""

        row = {
            "status": "to_be_updated",
            "ubm_target_date": "03.01.2026",
            "ubm_delivery_date": None,
            "status_flow": (
                '{"status":"to_be_issued","date":"01.01.2026"}\n'
                '{"status":"authority_approved","date":"02.01.2026"}'
            ),
            "cover_page_no": "CP-EDIT",
        }

        normalized = normalize_import_row(row, CompDoc)

        self.assertEqual(normalized["status_flow"][0]["date"], "03.01.2026")
        self.assertEqual(normalized["status_flow"][1]["status"], "authority_approved")
        self.assertEqual(normalized["status_flow"][2]["status"], "to_be_updated")

    def test_formula_like_user_text_is_stored_as_text(self):
        """Exported user text cannot become an active spreadsheet formula."""

        dataframe = pd.DataFrame([{"Name": '=HYPERLINK("https://invalid","open")'}])
        workbook = load_workbook(BytesIO(write_workbook(dataframe).getvalue()))
        worksheet = workbook["Compliance Documents"]

        self.assertEqual(worksheet["A2"].data_type, "s")
        self.assertTrue(worksheet["A2"].value.startswith("="))


class CompDocExcelExportLimitTests(TestCase):
    """Protect synchronous export memory with an explicit project-wide bound."""

    @override_settings(AWCENTER_MAX_COMPDOC_EXPORT_ROWS=0)
    def test_export_rejects_register_above_configured_limit(self):
        CompDoc.objects.create(name="Manual", cover_page_no="CP-LIMIT")
        CompDoc.objects.create(name="Manual 2", cover_page_no="CP-LIMIT-2")

        response = build_excel_response(CompDoc, CompDocSerializer)

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.data["code"], "COMPDOC_EXPORT_ROW_LIMIT")
