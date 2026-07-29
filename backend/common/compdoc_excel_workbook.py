"""Readable, safe, and re-importable CompDoc workbook presentation."""

from io import BytesIO

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from .compdoc_excel_dashboard import DATA_MAX_ROW, add_dashboard

SHEET_NAME = "Compliance Documents"
MAX_COLUMN_WIDTH = 50
WRAP_COLUMNS = {
    "Signature Panel",
    "Requirements",
    "Status Flow",
    "Tech Doc No",
    "Tech Doc Issue",
    "Delivered Tech Doc Issue",
    "Notes",
}
REFERENCE_VALUES = {
    "Status": (
        "to_be_issued",
        "airworthiness_review",
        "to_be_re-submitted",
        "to_be_updated",
        "authority_review",
        "authority_approved",
        "unknown",
    ),
    "Cat": ("1", "2", "3", "not_retained", "retained"),
    "Moc": tuple(str(value) for value in range(10)) + ("M",),
}
STATUS_COLORS = {
    "to_be_issued": "FFFECACA",
    "airworthiness_review": "FFFEF3C7",
    "to_be_re-submitted": "FFFED7AA",
    "to_be_updated": "FFFEF9C3",
    "authority_review": "FFBAE6FD",
    "authority_approved": "FFBBF7D0",
}


def write_workbook(dataframe):
    """Write an import-compatible register with a formula-driven dashboard."""

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=SHEET_NAME)
        workbook = writer.book
        worksheet = writer.sheets[SHEET_NAME]
        style_data_sheet(worksheet, tuple(dataframe.columns))
        add_dashboard(workbook, worksheet, tuple(dataframe.columns))
        workbook.properties.title = "AW Center Compliance Documents"
        workbook.properties.subject = "Editable register with operational dashboard"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    buffer.seek(0)
    return buffer


def style_data_sheet(worksheet, columns):
    """Apply navigation, validation, safety, and semantic status styling."""

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.sheet_properties.tabColor = "FF64748B"
    worksheet.row_dimensions[1].height = 30
    _style_header(worksheet)
    _style_body(worksheet, columns)
    _set_column_widths(worksheet)
    _add_table(worksheet)
    _add_status_rules(worksheet, columns)
    _add_validations(worksheet)
    _force_untrusted_text_cells(worksheet)
    worksheet.print_title_rows = "1:1"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0


def _style_header(worksheet):
    descriptions = {
        "Name": "Keep all column headers unchanged; they are the AW Center import contract.",
        "Status Flow": "One JSON event per line. Keep this column for lossless re-import.",
        "Status": "Changing this value appends a dated event during import.",
        "UBM Target Date": "Accepted formats include YYYY-MM-DD and DD.MM.YYYY.",
    }
    for cell in worksheet[1]:
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF16324F")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.value in descriptions:
            cell.comment = Comment(descriptions[cell.value], "AW Center")


def _style_body(worksheet, columns):
    for column_index, column in enumerate(columns, start=1):
        wrap = column in WRAP_COLUMNS
        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row_index, column_index)
            cell.font = Font(name="Aptos", size=10, color="FF1E293B")
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            worksheet.row_dimensions[row_index].height = 24


def _set_column_widths(worksheet):
    for column_index in range(1, worksheet.max_column + 1):
        cells = next(worksheet.iter_cols(min_col=column_index, max_col=column_index))
        values = (str(cell.value or "").splitlines() for cell in cells)
        longest = max((len(line) for lines in values for line in lines), default=0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(longest + 2, 12), MAX_COLUMN_WIDTH
        )


def _add_table(worksheet):
    if worksheet.max_column < 1 or worksheet.max_row < 2:
        return
    table = Table(displayName="ComplianceDocumentsTable", ref=worksheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _add_status_rules(worksheet, columns):
    if "Status" not in columns or worksheet.max_row < 2:
        return
    letter = get_column_letter(columns.index("Status") + 1)
    area = f"{letter}2:{letter}{worksheet.max_row}"
    for status, color in STATUS_COLORS.items():
        fill = PatternFill("solid", fgColor=color)
        worksheet.conditional_formatting.add(
            area,
            FormulaRule(formula=[f'${letter}2="{status}"'], fill=fill),
        )


def _add_validations(worksheet):
    for column_index, header in enumerate(worksheet[1], start=1):
        values = REFERENCE_VALUES.get(str(header.value))
        if not values:
            continue
        validation = DataValidation(
            type="list",
            formula1=f'"{",".join(values)}"',
            allow_blank=True,
            errorStyle="stop",
            errorTitle="Unsupported value",
            error="Choose a value from the AW Center import list.",
            showErrorMessage=True,
        )
        worksheet.add_data_validation(validation)
        letter = get_column_letter(column_index)
        validation.add(f"{letter}2:{letter}{DATA_MAX_ROW}")


def _force_untrusted_text_cells(worksheet):
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
                cell.data_type = "s"
