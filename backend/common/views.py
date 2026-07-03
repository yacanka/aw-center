from django.shortcuts import get_object_or_404
from django import forms
from django.http import HttpResponse

from rest_framework.viewsets import ModelViewSet
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from awcenter.pagination import StandardResultsSetPagination

from dateutil import parser
from datetime import datetime, date
import json
from io import BytesIO
from pprint import pprint

import math
from utils.arrays import safe_get
from common.compdoc_import import LIST_IMPORT_FIELDS, build_mapping_preview, choose_header_row, read_mapped_excel

PAGINATION_QUERY_PARAMETERS = {"page", "page_size"}
TEXT_FIELD_TYPES = {"CharField", "TextField", "EmailField"}


def get_query_values(request, name):
    """Return non-empty query values for DRF or Django requests."""

    query_parameters = getattr(request, "query_params", request.GET)
    values = query_parameters.getlist(name)
    return [value for value in values if value not in (None, "")]


def get_filter_expression(field, values):
    """Build a safe lookup expression for a model field and values."""

    if not values:
        return None
    field_type = field.get_internal_type()
    if len(values) > 1:
        return f"{field.name}__in", values
    if field_type in TEXT_FIELD_TYPES:
        return f"{field.name}__icontains", values[0]
    return field.name, values[0]


def filtered_queryset(request, queryset):
    """Apply safe server-side filters for model-backed list querysets."""

    model = getattr(queryset, "model", None)
    if model is None:
        return queryset

    fields = {field.name: field for field in model._meta.fields}
    for name, field in fields.items():
        if name in PAGINATION_QUERY_PARAMETERS:
            continue
        expression = get_filter_expression(field, get_query_values(request, name))
        if expression:
            queryset = queryset.filter(**{expression[0]: expression[1]})
    return queryset


def paginated_response(request, queryset, serializer_class):
    """Serialize a queryset using the standard paginated response contract."""

    queryset = filtered_queryset(request, queryset)
    paginator = StandardResultsSetPagination()
    page = paginator.paginate_queryset(queryset, request)
    serializer = serializer_class(page, many=True, context={"request": request})
    return paginator.get_paginated_response(serializer.data)

def history_view_set_factory(model, serializer_class, view_permission_classes):
    class DynamicHistoryViewSet(APIView):
        def get(self, request, pk):
            obj = get_object_or_404(model, pk=pk)
            obj_history = obj.history.all().order_by("-history_date", "-history_id")
            return paginated_response(request, obj_history, serializer_class)

        permission_classes = view_permission_classes

    return DynamicHistoryViewSet

def panel_view_set_factory(model, view_serializer_class, view_permission_classes):
    class DynamicPanelViewSet(ModelViewSet):
        permission_classes = view_permission_classes
        serializer_class = view_serializer_class
        queryset = model.objects.all()

    return DynamicPanelViewSet

def responsible_view_set_factory(model, view_serializer_class, view_permission_classes):
    class DynamicResponsibleViewSet(ModelViewSet):
        permission_classes = view_permission_classes
        serializer_class = view_serializer_class
        queryset = model.objects.all()

        def get_queryset(self):
            qs = model.objects.select_related("panel")

            panel = self.request.query_params.get("panel")
            if panel:
                qs = qs.filter(panel__name__iexact=panel)

            return qs

    return DynamicResponsibleViewSet

def view_set_factory(model, serializer_class, view_permission_classes):
    class DynamicViewSet(APIView):
        def get(self, request):
            objs = model.objects.all().order_by("-id")
            return paginated_response(request, objs, serializer_class)

        def post(self, request):
            serializer = serializer_class(data=request.data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        def delete(self, request):
            count, _ = model.objects.all().delete()
            return Response({"message": f"{count} records deleted successfully."}, status=status.HTTP_200_OK)

        permission_classes = view_permission_classes

    return DynamicViewSet

def view_set_obj_factory(model, serializer_class, view_permission_classes):
    class DynamicViewSet(APIView):
        def get(self, request, pk):
            obj = get_object_or_404(model, pk=pk)
            serializer = serializer_class(obj)
            return Response(serializer.data)

        def put(self, request, pk):
            obj = get_object_or_404(model, pk=pk)
            serializer = serializer_class(obj, data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        def patch(self, request, pk):
            obj = get_object_or_404(model, pk=pk)
            serializer = serializer_class(obj, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_200_OK)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        def delete(self, request, pk):
            obj = get_object_or_404(model, pk=pk)
            serializer = serializer_class(obj)
            obj.delete()
            return Response(serializer.data, status=status.HTTP_204_NO_CONTENT)

        permission_classes = view_permission_classes

    return DynamicViewSet

####################

wrap_style_cols = ["Signature Panel", "Requirements", "Status Flow", "Tech Doc No", "Tech Doc Issue", "Delivered Tech Doc Issue"]
list_cols = ["Signature Panel", "Requirements", "Status Flow"]

def excel_creator_factory(model, serializer_class, view_permission_classes):
    class ExcelCreator(APIView):
        def get(self, request):
            import pandas as pd
            from openpyxl.formatting.rule import FormulaRule
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            objs = model.objects.all()
            serializer = serializer_class(objs, many=True)
            df = pd.DataFrame(serializer.data)

            wrap_text_style = Alignment(wrap_text=True)

            sheet_name = "Compliance Documents"
            max_col_width = 50

            buffer = BytesIO()

            if "tech_doc_no_2" in df.columns:
                df["tech_doc_no"] += "\n" + df["tech_doc_no_2"]
                df.drop("tech_doc_no_2", axis=1, inplace=True)

            if "tech_doc_issue_2" in df.columns:
                df["tech_doc_issue"] += "\n" + df["tech_doc_issue_2"]
                df.drop("tech_doc_issue_2", axis=1, inplace=True)

            if "delivered_tech_doc_issue_2" in df.columns:
                df["delivered_tech_doc_issue"] += "\n" + df["delivered_tech_doc_issue_2"]
                df.drop("delivered_tech_doc_issue_2", axis=1, inplace=True)

            if "status_flow" in df.columns:
                print(type(df["status_flow"]))
                pprint(df["status_flow"])
                for flow in df["status_flow"]:
                    print(type(flow))
                    pprint(flow)

                df_exploded = df.explode("status_flow")
                df_exploded = pd.concat([df_exploded.drop(columns="status_flow"), df_exploded["status_flow"].apply(pd.Series)], axis=1)

                df["ubm_target_date"] = df["status_flow"].apply(lambda x: next((item.get("date") for item in x if item.get("status") == "to_be_issued"), None))
                df["ubm_delivery_date"] = df["status_flow"].apply(lambda x: next((item.get("date") for item in x if item.get("status") == "authority_review"), None))
                df["status"] = df["status_flow"].apply(lambda x: x[-1].get("status") if len(x) > 0 else None)

            df.drop(["id", "path", "created_time"], axis=1, inplace=True)

            df.columns = [str(column).replace('_', ' ').title() for column in df.columns]

            for i, column in enumerate(df.columns):
                if column in list_cols:
                    df[column] = df[column].apply(lambda x: "\n".join(map(str, x)) if isinstance(x, (list, tuple)) else ([] if (pd.isna(x)) else [x]))

            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                wb = writer.book
                ws = writer.sheets[sheet_name]

                for i, col in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(i)
                    max_len = len(str(col))
                    for cell in ws[col_letter]:
                        val = cell.value
                        if val is None:
                            l = 0
                        else:
                            if isinstance(val, (list, tuple)):
                                s = ", ".join(map(str, val))
                            else:
                                s = str(val)
                            l = len(s)
                        if l > max_len:
                            max_len = l

                    new_width = min(max_len + 2, max_col_width)
                    ws.column_dimensions[col_letter].width = new_width

                last_row = ws.max_row
                last_col_letter = get_column_letter(last_row)
                data_range = f"A{2}:{last_col_letter}{last_row}"

                fill = PatternFill(start_color="87CEB3", end_color="87CEB3", fill_type="solid")
                formula = ["MOD(ROW(),2)=0"]
                rule = FormulaRule(formula=formula, fill=fill)
                ws.conditional_formatting.add(data_range, rule)

                for i, column in enumerate(df.columns):
                    if column in wrap_style_cols:
                        letter = get_column_letter(i + 1)
                        for cell in ws[letter]:
                            cell.alignment = wrap_text_style

                for col_num, column_title in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = Font(bold=True, color="5BCEA8")
                    cell.fill = PatternFill(start_color="262626", end_color="262626", fill_type="solid")

            buffer.seek(0)

            res = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            res["Content-Disposition"] = 'attachment; filename="Compliance Documents.xlsx"'
            return res

        permission_classes = view_permission_classes

    return ExcelCreator

####################

reference_list = [
    "Name",
    "Panel",

    "Responsible",
    "Status",
    "Cat",
    "Moc",

    "Cover Page No",
    "Cover Page Issue",
    "Tech Doc No",
    "Tech Doc Issue",
]

def find_missing_elements(list_to_check, reference_list):
    missing_elements = [item for item in reference_list if item not in list_to_check]
    return missing_elements

class UploadForm(forms.Form):
    file = forms.FileField()

def upload_compdoc_factory(model, serializer_class, view_permission_classes):
    class UploadCompDoc(APIView):
        def post(self, request):
            form = UploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["file"]
                import pandas as pd

                model_fields = {field.name for field in model._meta.fields}
                header_result = choose_header_row(excel_file, pd, model_fields)
                if header_result.missing_fields:
                    return Response({
                        "message": "Missing column names exist.",
                        "missing_columns": header_result.missing_fields,
                    }, status=400)

                excel_file.seek(0)
                preview_df = pd.read_excel(excel_file, header=header_result.header_row_index)
                preview = build_mapping_preview(preview_df.columns, header_result)
                excel_file.seek(0)
                df = read_mapped_excel(excel_file, pd, header_result)
                df = df.astype(object).where(pd.notnull(df), None)

                for col in LIST_IMPORT_FIELDS:
                    if col not in df.columns:
                        df[col] = pd.Series([[] for _ in range(len(df))], dtype='object')

                for col in df.select_dtypes(include=["datetime64[ns]"]).columns:
                    df[col]= df[col].dt.date

                if request.query_params.get("preview") == "true":
                    invalid_serializer = []
                    for row_index, row in df.iterrows():
                        serializer = serializer_class(data=row.to_dict())
                        if not serializer.is_valid():
                            invalid_serializer.append({
                                "row": int(row_index) + 2 + header_result.header_row_index,
                                "name": row.get("name") or f"Row {int(row_index) + 1}",
                                "error": serializer.errors,
                            })
                    return Response({**preview, "invalid_documents": invalid_serializer}, status=200)

                if request.query_params.get("confirm_import") != "true":
                    return Response({"message": "Import preview confirmation is required."}, status=409)

                #df = df.applymap(lambda x: None if pd.isna(x) else x)
                #df = df.replace({pd.NA: None, np.nan: None})
                #df = df.fillna(value=None)

                try:
                    invalid_serializer = []
                    for _, row in df.iterrows():
                        try:
                            row["status"] = row["status"].strip().lower().replace('.', '').replace(' ', '_')

                            ata = row.get("ata")
                            if pd.notna(ata):
                                row["ata"] = ata.strip()

                            cover_page_issue = row.get("cover_page_issue")
                            if pd.notna(cover_page_issue):
                                row["cover_page_issue"] = int(float(str(cover_page_issue).strip()))

                            tech_doc_no = row.get("tech_doc_no")
                            if pd.notna(tech_doc_no):
                                tech_doc_numbers = [str(line).strip() for line in str(tech_doc_no).strip().splitlines()]
                                row["tech_doc_no"] = safe_get(tech_doc_numbers, 0)
                                row["tech_doc_no_2"] = safe_get(tech_doc_numbers, 1)

                            tech_doc_issue = row.get("tech_doc_issue")
                            if pd.notna(tech_doc_issue):
                                tech_doc_issues = [int(float(str(line).strip())) for line in str(tech_doc_issue).strip().splitlines()]
                                row["tech_doc_issue"] = safe_get(tech_doc_issues, 0)
                                row["tech_doc_issue_2"] = safe_get(tech_doc_issues, 1)

                            delivered_tech_doc_issue = row.get("delivered_tech_doc_issue")
                            if pd.notna(delivered_tech_doc_issue):
                                delivered_tech_doc_issues = [int(float(str(line).strip())) for line in str(delivered_tech_doc_issue).strip().splitlines()]
                                row["delivered_tech_doc_issue"] = safe_get(delivered_tech_doc_issues, 0)
                                row["delivered_tech_doc_issue_2"] = safe_get(delivered_tech_doc_issues, 1)

                            status_flow = [json.loads(line.replace("'", '"')) for line in row["status_flow"].strip().split("\n")] if row.get('status_flow') else []
                            if len(status_flow) == 0:
                                ubm_target_date = row.get("ubm_target_date")
                                target_initial_date = str(date.today().strftime("%d.%m.%Y"))
                                if pd.notna(ubm_target_date):
                                    if isinstance(ubm_target_date, str):
                                        target_initial_date = parser.parse(ubm_target_date).strftime("%d.%m.%Y")
                                    elif isinstance(ubm_target_date, (datetime, date)):
                                        target_initial_date = ubm_target_date.strftime("%d.%m.%Y")
                                status_flow.append({"status": "to_be_issued", "date": str(target_initial_date)})

                                ubm_delivery_date = row.get("ubm_delivery_date")
                                if pd.notna(ubm_delivery_date):
                                    if isinstance(ubm_delivery_date, str):
                                        delivery_initial_date = parser.parse(ubm_delivery_date).strftime("%d.%m.%Y")
                                    elif isinstance(ubm_delivery_date, (datetime, date)):
                                        delivery_initial_date = ubm_delivery_date.strftime("%d.%m.%Y")
                                    else:
                                        delivery_initial_date = str(date.today().strftime("%d.%m.%Y"))

                                    status_flow.append({"status": "authority_review", "date": str(delivery_initial_date)})

                                status = row.get("status")
                                if pd.notna(status) and status != "to_be_issued" and status != "authority_review":
                                    status_flow.append({"status": status, "date": str(date.today().strftime("%d.%m.%Y"))})

                            row["status_flow"] = status_flow

                            pprint(row.apply(type))
                            serializer = serializer_class(data=row.to_dict())
                            if serializer.is_valid():
                                serializer.save()
                            else:
                                invalid_serializer.append({"name": row["name"], "error": serializer.errors})
                        except Exception as e:
                            print(row["name"], e)
                            invalid_serializer.append({"name": row["name"], "error": str(e)})
                    if invalid_serializer:
                        return Response({"message": "Uploaded successfully. But some documents can not imported.", "invalid_documents": invalid_serializer}, status=201)
                    return Response({"message": "Uploaded successfully. Database updated."}, status=201)
                except Exception as e:
                    return Response({"message": f"Can not upload excel: {e}"}, status=400)
            else:
                return Response({"message": "Not a valid form."}, status=400)

        permission_classes = view_permission_classes

    return UploadCompDoc