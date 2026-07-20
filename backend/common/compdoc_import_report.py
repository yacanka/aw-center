"""Safe Excel remediation reports for CompDoc import audits."""

import json
import re
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ILLEGAL_CELL_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
FORMULA_PREFIXES = ("=", "+", "-", "@")
MAX_CELL_LENGTH = 2_000


def build_import_audit_report(audit):
    """Return a downloadable workbook containing sanitized remediation evidence."""

    workbook = Workbook()
    add_summary_sheet(workbook.active, audit)
    add_error_sheet(workbook.create_sheet("Rejected Rows"), audit.error_summary)
    add_mapping_sheet(workbook.create_sheet("Column Mappings"), audit)
    response = HttpResponse(workbook_bytes(workbook), content_type=XLSX_CONTENT_TYPE)
    filename = f"compdoc-import-audit-{audit.pk}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def add_summary_sheet(worksheet, audit):
    """Write audit identity, outcome, and source integrity evidence."""

    worksheet.title = "Summary"
    append_safe_rows(worksheet, [("Field", "Value"), *summary_rows(audit)])
    style_worksheet(worksheet)


def summary_rows(audit):
    """Return ordered, sanitized audit summary values."""

    return [
        ("Audit ID", audit.pk),
        ("Project", audit.project_slug),
        ("Source file", audit.source_filename),
        ("Source SHA-256", audit.source_sha256),
        ("Imported by", audit.imported_by_username),
        ("Status", audit.status),
        ("Started", format_datetime(audit.started_at)),
        ("Completed", format_datetime(audit.completed_at)),
        ("Duration (ms)", audit.duration_ms),
        ("Header row", audit.header_row),
        ("Total rows", audit.total_rows),
        ("Created", audit.created_count),
        ("Updated", audit.updated_count),
        ("Unchanged", audit.unchanged_count),
        ("Rejected", audit.rejected_count),
        ("Request reference", audit.request_id),
    ]


def add_error_sheet(worksheet, errors):
    """Write bounded rejected-row evidence and actionable recovery guidance."""

    header = ("Row", "Document", "Code", "Problem", "Affected fields", "Suggested action")
    rows = [
        (
            error.get("row"),
            error.get("name", ""),
            error.get("code", ""),
            error.get("detail", ""),
            format_fields(error.get("fields")),
            recovery_action(error.get("code")),
        )
        for error in errors
        if isinstance(error, dict)
    ]
    append_safe_rows(worksheet, [header, *rows])
    style_worksheet(worksheet, wrap_columns={4, 5, 6})


def add_mapping_sheet(worksheet, audit):
    """Write detected, missing, and unmapped column decisions."""

    rows = [("Source column", "Model field", "State")]
    rows.extend((item.get("source"), item.get("target"), "Mapped") for item in audit.mapped_columns)
    rows.extend(("", value, "Missing required column") for value in audit.missing_columns)
    rows.extend((value, "", "Unmapped source column") for value in audit.unmapped_columns)
    append_safe_rows(worksheet, rows)
    style_worksheet(worksheet)


def append_safe_rows(worksheet, rows):
    """Append rows after neutralizing formulas and illegal XML characters."""

    for row in rows:
        worksheet.append([safe_cell(value) for value in row])


def safe_cell(value):
    """Return an XLSX-safe scalar without executable formula semantics."""

    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = ILLEGAL_CELL_CHARACTERS.sub("", str(value))[:MAX_CELL_LENGTH]
    return f"'{text}" if text.startswith(FORMULA_PREFIXES) else text


def format_fields(fields):
    """Return deterministic bounded field-validation evidence."""

    if not isinstance(fields, dict) or not fields:
        return ""
    return json.dumps(fields, ensure_ascii=False, sort_keys=True)


def recovery_action(code):
    """Map stable import codes to concise remediation guidance."""

    actions = {
        "ROW_VALIDATION_FAILED": "Correct the affected fields in the source row and import again.",
        "ROW_TRANSFORM_FAILED": "Correct the row value formats and import again.",
        "ROW_SAVE_FAILED": "Retry the import; use the request reference if the error continues.",
        "ROW_DUPLICATE_KEY": "Keep one row for each cover page number and preview again.",
        "COMPDOC_IMPORT_MISSING_COLUMNS": "Add the missing required columns and import again.",
        "COMPDOC_IMPORT_ROW_LIMIT": "Split the workbook into smaller files and import each file.",
    }
    return actions.get(str(code), "Review the problem and source row, then import again.")


def format_datetime(value):
    """Render an aware timestamp in the active application timezone."""

    return timezone.localtime(value).isoformat() if value else ""


def style_worksheet(worksheet, wrap_columns=None):
    """Apply readable, bounded styling to one generated worksheet."""

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="263238")
    for column in wrap_columns or set():
        for cell in worksheet[get_column_letter(column)]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    set_column_widths(worksheet)


def set_column_widths(worksheet):
    """Autosize columns while preventing unbounded workbook dimensions."""

    for index in range(1, worksheet.max_column + 1):
        letter = get_column_letter(index)
        width = max((len(str(cell.value or "")) for cell in worksheet[letter]), default=0)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), 60)


def workbook_bytes(workbook):
    """Serialize one workbook into an in-memory byte string."""

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
