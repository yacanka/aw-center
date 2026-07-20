from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from .models import CompDocImportAudit

User = get_user_model()


class CompDocImportReportTests(TestCase):
    """Verify secure, permission-protected import remediation workbooks."""

    def setUp(self):
        """Create viewers and formula-like sanitized audit evidence."""

        self.viewer = User.objects.create_user("audit-viewer", password="StrongPass!123")
        self.viewer.user_permissions.add(
            Permission.objects.get(codename="view_compdocimportaudit")
        )
        self.ordinary = User.objects.create_user("ordinary-user", password="StrongPass!123")
        self.audit = self.create_audit()

    def create_audit(self):
        """Persist representative formula-like audit evidence."""

        return CompDocImportAudit.objects.create(
            project_slug="ozgur",
            source_filename='=HYPERLINK("https://invalid.example")',
            source_size=128,
            source_sha256="a" * 64,
            imported_by=self.viewer,
            imported_by_username=self.viewer.username,
            request_id="request-reference",
            mapped_columns=[{"source": "=Source", "target": "name"}],
            missing_columns=["@required"],
            unmapped_columns=["-unused"],
            total_rows=2,
            rejected_count=1,
            error_summary=[formula_like_error()],
            status=CompDocImportAudit.Status.FAILED,
        )

    def test_authorized_report_contains_actionable_sanitized_evidence(self):
        """The workbook is useful without copying original workbook contents."""

        client = APIClient()
        client.force_authenticate(self.viewer)
        response = client.get(self.report_url)
        workbook = load_workbook(BytesIO(response.content), data_only=False)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn(str(self.audit.pk), response["Content-Disposition"])
        self.assertEqual(workbook.sheetnames, ["Summary", "Rejected Rows", "Column Mappings"])
        self.assert_report_values(workbook)
        self.assert_formula_cells_are_text(workbook)
        self.assertNotIn("secret-original-cell", all_cell_text(workbook))

    def test_report_requires_explicit_audit_permission(self):
        """Authentication alone never exposes report evidence."""

        anonymous = APIClient().get(self.report_url)
        client = APIClient()
        client.force_authenticate(self.ordinary)
        denied = client.get(self.report_url)

        self.assertEqual(anonymous.status_code, 401)
        self.assertEqual(denied.status_code, 403)

    def assert_report_values(self, workbook):
        """Assert summary, row guidance, and mapping decisions are present."""

        summary = dict(workbook["Summary"].iter_rows(min_row=2, values_only=True))
        errors = workbook["Rejected Rows"]
        mappings = workbook["Column Mappings"]
        self.assertEqual(summary["Audit ID"], str(self.audit.pk))
        self.assertEqual(summary["Rejected"], 1)
        self.assertEqual(summary["Unchanged"], 0)
        self.assertEqual(errors["A2"].value, 3)
        self.assertIn("Correct the affected fields", errors["F2"].value)
        self.assertEqual(mappings["C2"].value, "Mapped")
        self.assertEqual(mappings["C3"].value, "Missing required column")
        self.assertEqual(mappings["C4"].value, "Unmapped source column")

    def assert_formula_cells_are_text(self, workbook):
        """Assert user-controlled leading formula characters are neutralized."""

        cells = [
            workbook["Summary"]["B4"],
            workbook["Rejected Rows"]["B2"],
            workbook["Rejected Rows"]["D2"],
            workbook["Column Mappings"]["A2"],
            workbook["Column Mappings"]["B3"],
            workbook["Column Mappings"]["A4"],
        ]
        for cell in cells:
            self.assertTrue(cell.value.startswith("'"))
            self.assertEqual(cell.data_type, "s")

    @property
    def report_url(self):
        """Return the audited report endpoint for the test record."""

        return f"/projects/import-audits/{self.audit.pk}/report/"


def all_cell_text(workbook):
    """Return searchable text from every generated report cell."""

    return " ".join(
        str(cell.value or "")
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
    )


def formula_like_error():
    """Return an audit error containing spreadsheet formula prefixes."""

    return {
        "row": 3,
        "name": "=Dangerous document",
        "code": "ROW_VALIDATION_FAILED",
        "detail": "+Correct the invalid value",
        "fields": {"status": ["-Required"]},
    }
