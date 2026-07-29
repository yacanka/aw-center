"""Formula-driven dashboard presentation for CompDoc exports."""

from collections import Counter

from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DASHBOARD_SHEET_NAME = "Dashboard"
DATA_MAX_ROW = 10001
NAVY, SLATE = "FF16324F", "FF334155"
LIGHT = "FFF8FAFC"
WHITE = "FFFFFFFF"
CARD_COLORS = ("FF1D4ED8", "FF15803D", "FF0369A1", "FFB45309", "FF0F766E")
STATUS_METADATA = (
    ("To be issued", "to_be_issued", "FFFECACA"), ("Airworthiness review", "airworthiness_review", "FFFEF3C7"),
    ("To be re-submitted", "to_be_re-submitted", "FFFED7AA"), ("To be updated", "to_be_updated", "FFFEF9C3"),
    ("Authority review", "authority_review", "FFBAE6FD"), ("Authority approved", "authority_approved", "FFBBF7D0"),
    ("Unknown", "unknown", "FFE2E8F0")
)

def add_dashboard(workbook, data_sheet, columns):
    """Add a professional summary sheet backed by editable register formulas."""
    dashboard = workbook.create_sheet(DASHBOARD_SHEET_NAME)
    dashboard.sheet_view.showGridLines = False
    dashboard.sheet_view.zoomScale = 90
    dashboard.sheet_properties.tabColor = "FF1D4ED8"
    _set_dimensions(dashboard)
    _add_title(dashboard)
    ranges = _data_ranges(columns)
    _add_kpis(dashboard, ranges)
    _add_status_section(dashboard, ranges["status"])
    _add_panel_section(dashboard, data_sheet, ranges["panel"])
    _add_guide(dashboard)
    dashboard.freeze_panes = "A5"
    dashboard.print_area = "A1:O28"
    dashboard.sheet_properties.pageSetUpPr.fitToPage = True
    dashboard.page_setup.fitToWidth = 1
    dashboard.page_setup.fitToHeight = 1
    workbook.active = dashboard
    return dashboard

def _data_ranges(columns):
    def column_range(header):
        if header not in columns:
            return None
        letter = get_column_letter(columns.index(header) + 1)
        return f"'Compliance Documents'!${letter}$2:${letter}${DATA_MAX_ROW}"

    return {
        "name": column_range("Name"),
        "panel": column_range("Panel"),
        "status": column_range("Status")
    }

def _set_dimensions(sheet):
    widths = {1: 22, 2: 14, 9: 20, 10: 14}
    for column in range(1, 16):
        sheet.column_dimensions[get_column_letter(column)].width = widths.get(column, 12)
    for row, height in {1: 28, 2: 28, 3: 22, 5: 24, 6: 28, 7: 28}.items():
        sheet.row_dimensions[row].height = height

def _add_title(sheet):
    sheet.merge_cells("A1:O2")
    title = sheet["A1"]
    title.value = "AW CENTER · COMPLIANCE DOCUMENT REGISTER"
    title.font = Font(name="Aptos Display", size=22, bold=True, color=WHITE)
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("A3:O3")
    subtitle = sheet["A3"]
    subtitle.value = "Live summary · Edit and re-import the Compliance Documents sheet"
    subtitle.font = Font(name="Aptos", size=10, color="FF475569")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

def _add_kpis(sheet, ranges):
    total = f'=COUNTA({ranges["name"]})' if ranges["name"] else "=0"
    approved = f'=COUNTIF({ranges["status"]},"authority_approved")' if ranges["status"] else "=0"
    review = f'=COUNTIF({ranges["status"]},"authority_review")' if ranges["status"] else "=0"
    cards = (
        ("A", "C", "Total documents", total, "#,##0"),
        ("D", "F", "Authority approved", approved, "#,##0"),
        ("G", "I", "Authority review", review, "#,##0"),
        ("J", "L", "Open workflow", "=MAX(A6-D6,0)", "#,##0"),
        ("M", "O", "Completion rate", "=IF(A6=0,0,D6/A6)", "0%"),
    )
    for index, card in enumerate(cards):
        _add_kpi_card(sheet, index, card)

def _add_kpi_card(sheet, index, card):
    start, end, label, formula, number_format = card
    sheet.merge_cells(f"{start}5:{end}5")
    sheet.merge_cells(f"{start}6:{end}7")
    label_cell, value_cell = sheet[f"{start}5"], sheet[f"{start}6"]
    label_cell.value, value_cell.value = label, formula
    for cell in (label_cell, value_cell):
        cell.fill = PatternFill("solid", fgColor=CARD_COLORS[index])
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = Border(bottom=Side(style="medium", color=WHITE))
    value_cell.font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    value_cell.number_format = number_format

def _add_status_section(sheet, status_range):
    _table_header(sheet, 1, 2, "Status breakdown")
    for row, (label, status, color) in enumerate(STATUS_METADATA, start=11):
        sheet.cell(row, 1, label)
        formula = f'=COUNTIF({status_range},"{status}")' if status_range else "=0"
        sheet.cell(row, 2, formula)
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=color)
        sheet.cell(row, 2).number_format = "#,##0"
    chart = DoughnutChart()
    chart.title = "Lifecycle status"
    chart.holeSize = 58
    chart.varyColors = True
    chart.legend.position = "r"
    chart.add_data(Reference(sheet, min_col=2, min_row=10, max_row=17), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=1, min_row=11, max_row=17))
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True
    chart.height = 7.2
    chart.width = 10.5
    sheet.add_chart(chart, "D10")

def _add_panel_section(sheet, data_sheet, panel_range):
    _table_header(sheet, 9, 10, "Panel workload")
    panels = _top_panels(data_sheet)
    if not panel_range or not panels:
        _add_empty_panel_message(sheet)
        return
    for row, panel in enumerate(panels, start=11):
        label = sheet.cell(row, 9)
        label.value = panel
        label.data_type = "s"
        escaped = panel.replace('"', '""')
        sheet.cell(row, 10, f'=COUNTIF({panel_range},"{escaped}")')
        sheet.cell(row, 10).number_format = "#,##0"
    _add_panel_chart(sheet, len(panels))

def _add_empty_panel_message(sheet):
    sheet.merge_cells("I11:O13")
    sheet["I11"] = "No panel data is available in this export."
    sheet["I11"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["I11"].fill = PatternFill("solid", fgColor=LIGHT)

def _add_panel_chart(sheet, panel_count):
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Documents by panel"
    chart.legend = None
    chart.y_axis.majorUnit = 1
    chart.y_axis.numFmt = "0"
    chart.add_data(Reference(sheet, min_col=10, min_row=10, max_row=10 + panel_count), titles_from_data=True)
    chart.set_categories(Reference(sheet, min_col=9, min_row=11, max_row=10 + panel_count))
    chart.height = 7.2
    chart.width = 10.5
    sheet.add_chart(chart, "K10")

def _top_panels(data_sheet):
    headers = {cell.value: cell.column for cell in data_sheet[1]}
    column = headers.get("Panel")
    if not column:
        return []
    counts = Counter(
        str(data_sheet.cell(row, column).value).strip()
        for row in range(2, data_sheet.max_row + 1)
        if data_sheet.cell(row, column).value
    )
    return [panel for panel, _count in counts.most_common(8)]

def _section_title(sheet, area, title):
    sheet.merge_cells(area)
    cell = sheet[area.split(":")[0]]
    cell.value = title
    cell.fill = PatternFill("solid", fgColor=SLATE)
    cell.font = Font(name="Aptos", bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def _table_header(sheet, start_column, end_column, title):
    for column in range(start_column, end_column + 1):
        cell = sheet.cell(10, column)
        cell.fill = PatternFill("solid", fgColor=SLATE)
        cell.font = Font(name="Aptos", bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.cell(10, start_column, title)
    sheet.cell(10, end_column, "Documents")

def _add_guide(sheet):
    _section_title(sheet, "A25:O25", "Using this workbook")
    sheet.merge_cells("A26:O28")
    cell = sheet["A26"]
    cell.value = (
        "Use Dashboard for monitoring. Edit only the Compliance Documents sheet, keep its headers "
        "unchanged, then import that workbook back into AW Center. Dashboard formulas refresh in Excel."
    )
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    cell.font = Font(name="Aptos", size=10, color="FF334155")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
