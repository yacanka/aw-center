"""Characterization tests for Word comparison helpers."""

from io import BytesIO

from django.test import SimpleTestCase
from openpyxl import load_workbook

from word.service.compare import (
    align_paragraphs_indexed,
    auto_thresholds,
    ratio,
    split_sentences,
    tokenize_words,
    u_normalize,
)
from word.views import write_excel_report_openpyxl


class WordComparisonHelperTests(SimpleTestCase):
    """Protect legacy Word comparison behavior during view decomposition."""

    def test_normalize_removes_zero_width_and_collapses_whitespace(self):
        self.assertEqual(u_normalize("  A\u00A0\u200B  B\r\n "), "A B")

    def test_split_sentences_packs_short_sentences(self):
        self.assertEqual(split_sentences("One. Two? Three!"), ["One. Two? Three!"])
        self.assertEqual(split_sentences(""), [])

    def test_tokenize_words_keeps_punctuation_tokens(self):
        self.assertEqual(tokenize_words("Hello, AW!"), ["Hello", ",", "AW", "!"])

    def test_ratio_uses_sequence_matcher_similarity(self):
        self.assertEqual(ratio("same", "same"), 1.0)
        self.assertLess(ratio("abc", "xyz"), 0.5)

    def test_auto_thresholds_return_legacy_defaults_for_empty_inputs(self):
        self.assertEqual(auto_thresholds([], []), (0.92, 0.86))

    def test_align_paragraphs_keeps_equal_insert_delete_and_replace_shapes(self):
        first_lines = [(0, "same"), (1, "old text"), (2, "removed")]
        second_lines = [(0, "same"), (1, "new text"), (2, "added")]
        aligned = align_paragraphs_indexed(first_lines, second_lines, 0.99, 0.50)
        self.assertEqual(aligned[0], (0, "same", 0, "same", "equal"))
        self.assertIn((1, "old text", 1, "new text", "replace"), aligned)

    def test_excel_report_neutralizes_untrusted_document_text(self):
        output = BytesIO()

        write_excel_report_openpyxl(
            output,
            [
                {
                    "Tag": "replace",
                    "A_Index": 1,
                    "B_Index": 1,
                    "A_Text": "=1+1",
                    "B_Text": "@SUM(A1:A2)",
                    "A_Len": 4,
                    "B_Len": 11,
                    "Similarity": 0.5,
                }
            ],
            {"equal": 0, "insert": 0, "delete": 0, "replace": 1},
        )

        worksheet = load_workbook(BytesIO(output.getvalue()), data_only=False)["Diff"]
        headers = {cell.value: cell.column for cell in worksheet[1]}
        for field in ("A_Text", "B_Text"):
            cell = worksheet.cell(row=2, column=headers[field])
            self.assertNotEqual(cell.data_type, "f")
            self.assertTrue(cell.value.startswith("'"))
