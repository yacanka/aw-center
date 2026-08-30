"""Acceptance tests for the canonical compliance aggregate."""

from datetime import date
from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import connection
from django.test import TestCase, override_settings
from openpyxl import load_workbook
from rest_framework.test import APIClient

from orgs.models import Panel, Project, ProjectRoleAssignment

from .models import (
    ComplianceDocument,
    CoverPage,
    NotificationPolicy,
    ReviewTask,
    TrackingProfile,
    WorkflowEvent,
)


class ComplianceApiTests(TestCase):
    def setUp(self):
        self.project = Project.objects.get(slug="ozgur")
        self.panel = Panel.objects.create(
            project=self.project,
            name="Flight",
            discipline="Systems",
            ata="27-00",
        )
        self.viewer = get_user_model().objects.create_user("viewer", password="safe-pass")
        self.editor = get_user_model().objects.create_user("editor", password="safe-pass")
        self.manager = get_user_model().objects.create_user("manager", password="safe-pass")
        self.assignee = get_user_model().objects.create_user("assignee", password="safe-pass")
        for user, role in (
            (self.viewer, ProjectRoleAssignment.Role.VIEWER),
            (self.editor, ProjectRoleAssignment.Role.EDITOR),
            (self.manager, ProjectRoleAssignment.Role.MANAGER),
            (self.assignee, ProjectRoleAssignment.Role.VIEWER),
        ):
            ProjectRoleAssignment.objects.create(
                project=self.project,
                domain=ProjectRoleAssignment.Domain.COMPLIANCE,
                role=role,
                user=user,
            )
        self.client = APIClient()

    @property
    def collection_url(self):
        return "/api/projects/ozgur/compliance-documents/"

    def create_document(self, *, owner=None):
        cover = CoverPage.objects.create(
            project=self.project,
            number="CP-001",
            issue="A",
        )
        return ComplianceDocument.objects.create(
            project=self.project,
            panel=self.panel,
            cover_page=cover,
            name="Flight Controls Compliance",
            tech_doc_no="TD-001",
            owner=owner or self.editor,
        )

    def test_viewer_cannot_create_but_editor_can(self):
        payload = {
            "panel": self.panel.pk,
            "cover_page": {"number": "CP-API", "issue": "A"},
            "name": "API-created document",
            "tech_doc_no": "TD-API",
        }
        self.client.force_authenticate(self.viewer)
        rejected = self.client.post(self.collection_url, payload, format="json")

        self.client.force_authenticate(self.editor)
        accepted = self.client.post(self.collection_url, payload, format="json")

        self.assertEqual(rejected.status_code, 403)
        self.assertEqual(accepted.status_code, 201)
        self.assertEqual(accepted.data["version"], 1)
        self.assertEqual(accepted.data["project_slug"], "ozgur")

    def test_optimistic_update_rejects_stale_version(self):
        document = self.create_document()
        url = f"{self.collection_url}{document.pk}/"
        self.client.force_authenticate(self.editor)

        updated = self.client.patch(
            url,
            {"version": 1, "notes": "Current value"},
            format="json",
        )
        stale = self.client.patch(
            url,
            {"version": 1, "notes": "Stale value"},
            format="json",
        )

        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.data["version"], 2)
        self.assertEqual(stale.status_code, 409)
        self.assertEqual(stale.data["code"], "VERSION_CONFLICT")
        document.refresh_from_db()
        self.assertEqual(document.notes, "Current value")

    def test_noop_document_update_does_not_advance_version(self):
        document = self.create_document()
        self.client.force_authenticate(self.editor)

        response = self.client.patch(
            f"{self.collection_url}{document.pk}/",
            {"version": document.version, "name": document.name},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["version"], document.version)

    def test_shared_cover_page_issue_has_an_independent_optimistic_version(self):
        first = self.create_document()
        second = ComplianceDocument.objects.create(
            project=self.project,
            panel=self.panel,
            cover_page=first.cover_page,
            name="Second shared-cover document",
            tech_doc_no="TD-SHARED",
        )
        self.client.force_authenticate(self.editor)

        changed = self.client.patch(
            f"{self.collection_url}{first.pk}/",
            {
                "version": first.version,
                "cover_page": {
                    "number": first.cover_page.number,
                    "issue": "B",
                    "version": first.cover_page.version,
                },
            },
            format="json",
        )
        stale = self.client.patch(
            f"{self.collection_url}{second.pk}/",
            {
                "version": second.version,
                "cover_page": {
                    "number": first.cover_page.number,
                    "issue": "C",
                    "version": 1,
                },
            },
            format="json",
        )

        self.assertEqual(changed.status_code, 200)
        self.assertEqual(changed.data["cover_page"]["issue"], "B")
        self.assertEqual(changed.data["cover_page"]["version"], 2)
        self.assertEqual(stale.status_code, 409)
        self.assertEqual(stale.data["code"], "COVER_PAGE_VERSION_CONFLICT")

    def test_table_filters_and_ordering_are_applied_by_the_server(self):
        document = self.create_document()
        ComplianceDocument.objects.create(
            project=self.project,
            panel=self.panel,
            cover_page=document.cover_page,
            name="Alpha secondary document",
            tech_doc_no="TD-002",
        )
        self.client.force_authenticate(self.viewer)

        filtered = self.client.get(self.collection_url, {"name": "secondary"})
        ordered = self.client.get(self.collection_url, {"ordering": "name"})
        invalid = self.client.get(self.collection_url, {"ordering": "owner"})

        self.assertEqual(filtered.status_code, 200)
        self.assertEqual(filtered.data["count"], 1)
        self.assertEqual(filtered.data["results"][0]["name"], "Alpha secondary document")
        self.assertEqual(ordered.status_code, 200)
        self.assertEqual(
            [row["name"] for row in ordered.data["results"]],
            ["Alpha secondary document", "Flight Controls Compliance"],
        )
        self.assertEqual(invalid.status_code, 400)

    def test_field_contract_reports_real_filter_and_sort_capabilities(self):
        self.client.force_authenticate(self.viewer)

        response = self.client.get(f"{self.collection_url}fields/")

        fields = {field["key"]: field for field in response.data["fields"]}
        self.assertEqual(response.data["schema_version"], 3)
        self.assertEqual(fields["name"]["filter_kind"], "text")
        self.assertTrue(fields["name"]["sortable"])
        self.assertEqual(fields["panel"]["option_source"], "panels")
        self.assertTrue(fields["status"]["choices"])

    def test_tracking_read_does_not_create_domain_state(self):
        document = self.create_document()
        self.client.force_authenticate(self.viewer)

        response = self.client.get(f"{self.collection_url}{document.pk}/tracking/")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["responsible_mode"], "automatic")
        self.assertFalse(response.data["notification_enabled"])
        self.assertFalse(TrackingProfile.objects.filter(document=document).exists())

    @patch("compliance.services.search_document_issue", return_value=(2, None))
    def test_docproof_refresh_persists_versioned_revision_evidence(self, search_issue):
        document = self.create_document()
        document.tech_doc_issue = "1"
        document.save(update_fields=["tech_doc_issue"])
        self.client.force_authenticate(self.editor)
        url = f"{self.collection_url}{document.pk}/tracking/docproof/"

        refreshed = self.client.post(url, {"version": 0}, format="json")
        stale = self.client.post(url, {"version": 0}, format="json")

        self.assertEqual(refreshed.status_code, 200)
        self.assertEqual(refreshed.data["docproof_status"], "revision_available")
        self.assertEqual(refreshed.data["docproof_issue"], "2")
        self.assertEqual(refreshed.data["version"], 1)
        self.assertEqual(stale.status_code, 409)
        self.assertEqual(search_issue.call_count, 2)

    def test_export_neutralizes_untrusted_formula_values(self):
        document = self.create_document()
        document.name = "=1+1"
        document.save(update_fields=["name"])
        self.client.force_authenticate(self.viewer)

        response = self.client.get(f"{self.collection_url}export/")

        self.assertEqual(response.status_code, 200)
        worksheet = load_workbook(BytesIO(response.content), data_only=False).active
        headers = {cell.value: cell.column for cell in worksheet[1]}
        name_cell = worksheet.cell(row=2, column=headers["name"])
        self.assertEqual(name_cell.data_type, "s")
        self.assertEqual(name_cell.value, "'=1+1")

    @override_settings(AWCENTER_MAX_COMPDOC_EXPORT_ROWS=1)
    def test_export_fails_closed_above_configured_row_limit(self):
        self.create_document()
        ComplianceDocument.objects.create(
            project=self.project,
            panel=self.panel,
            cover_page=CoverPage.objects.create(
                project=self.project,
                number="CP-002",
                issue="A",
            ),
            name="Second document",
        )
        self.client.force_authenticate(self.viewer)

        response = self.client.get(f"{self.collection_url}export/")

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.data["code"], "COMPDOC_EXPORT_ROW_LIMIT")

    def test_transition_event_and_projection_commit_together(self):
        document = self.create_document()
        self.client.force_authenticate(self.editor)

        response = self.client.post(
            f"{self.collection_url}{document.pk}/transitions/",
            {
                "version": 1,
                "status": "to_be_issued",
                "effective_date": "2026-08-11",
                "next_action_due_date": "2026-08-20",
                "reason": "Ready for issue",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        document.refresh_from_db()
        event = WorkflowEvent.objects.get(document=document)
        self.assertEqual(document.status, event.status)
        self.assertEqual(document.ubm_target_date, date(2026, 8, 11))
        self.assertEqual(document.version, 2)

    def test_only_manager_can_archive_and_restore(self):
        document = self.create_document()
        url = f"{self.collection_url}{document.pk}/archive/"
        self.client.force_authenticate(self.editor)
        rejected = self.client.post(url, {"version": 1, "reason": "Obsolete"}, format="json")

        self.client.force_authenticate(self.manager)
        archived = self.client.post(url, {"version": 1, "reason": "Obsolete"}, format="json")
        restored = self.client.post(
            f"{self.collection_url}{document.pk}/restore/",
            {"version": 2, "reason": "Required again"},
            format="json",
        )

        self.assertEqual(rejected.status_code, 403)
        self.assertFalse(archived.data["is_archived"] is False)
        self.assertFalse(restored.data["is_archived"])

    def test_group_editor_role_is_effective(self):
        group_user = get_user_model().objects.create_user("group-editor")
        group = Group.objects.create(name="Compliance Editors")
        group.user_set.add(group_user)
        ProjectRoleAssignment.objects.create(
            project=self.project,
            domain=ProjectRoleAssignment.Domain.COMPLIANCE,
            role=ProjectRoleAssignment.Role.EDITOR,
            group=group,
        )
        self.client.force_authenticate(group_user)

        response = self.client.post(
            self.collection_url,
            {
                "cover_page": {"number": "CP-GROUP", "issue": "A"},
                "name": "Group-created document",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)

    def test_review_is_version_bound_and_assignee_scoped(self):
        document = self.create_document()
        self.client.force_authenticate(self.editor)
        created = self.client.post(
            f"{self.collection_url}{document.pk}/reviews/",
            {
                "version": 1,
                "kind": "review",
                "assignee": self.assignee.pk,
                "request_note": "Please review this evidence.",
            },
            format="json",
        )
        task = ReviewTask.objects.get(pk=created.data["id"])

        self.client.force_authenticate(self.viewer)
        rejected = self.client.post(
            f"{self.collection_url}{document.pk}/reviews/{task.pk}/decision/",
            {"status": "approved", "decision_note": "Looks correct."},
            format="json",
        )
        self.client.force_authenticate(self.assignee)
        accepted = self.client.post(
            f"{self.collection_url}{document.pk}/reviews/{task.pk}/decision/",
            {"status": "approved", "decision_note": "Looks correct."},
            format="json",
        )

        self.assertEqual(created.status_code, 201)
        self.assertEqual(task.source_version, 1)
        self.assertEqual(rejected.status_code, 403)
        self.assertEqual(accepted.status_code, 200)

    def test_document_change_supersedes_pending_review(self):
        document = self.create_document()
        task = ReviewTask.objects.create(
            document=document,
            kind=ReviewTask.Kind.REVIEW,
            assignee=self.assignee,
            assignee_username=self.assignee.username,
            requested_by=self.editor,
            requested_by_username=self.editor.username,
            request_note="Review the current version.",
            source_version=document.version,
        )
        self.client.force_authenticate(self.editor)

        response = self.client.patch(
            f"{self.collection_url}{document.pk}/",
            {"version": document.version, "notes": "Changed after review"},
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        task.refresh_from_db()
        self.assertEqual(task.status, ReviewTask.Status.SUPERSEDED)

    def test_noop_work_update_does_not_advance_version(self):
        document = self.create_document(owner=self.editor)
        self.client.force_authenticate(self.editor)

        response = self.client.put(
            f"{self.collection_url}{document.pk}/work/",
            {
                "version": document.version,
                "owner": self.editor.pk,
                "owner_group": None,
                "next_action_due_date": None,
                "reason": "No changes",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data["version"], document.version)

    def test_notification_policy_revisions_serialize_on_project_row(self):
        self.client.force_authenticate(self.manager)
        url = f"{self.collection_url}notification-policy/"
        payload = {
            "version": 0,
            "event_rules": {"overdue": {"enabled": False}},
            "change_note": "Disable overdue notifications",
        }
        executed_queries = []

        def record_query(execute, sql, params, many, context):
            executed_queries.append(" ".join(sql.upper().split()))
            return execute(sql, params, many, context)

        with connection.execute_wrapper(record_query):
            first = self.client.put(url, payload, format="json")
        second = self.client.put(
            url,
            {
                **payload,
                "version": 1,
                "change_note": "Keep overdue notifications disabled",
            },
            format="json",
        )

        self.assertEqual(first.status_code, 200)
        self.assertEqual(first.data["version"], 1)
        self.assertEqual(second.status_code, 200)
        self.assertEqual(second.data["version"], 2)
        project_policies = NotificationPolicy.objects.filter(project=self.project)
        self.assertEqual(project_policies.count(), 2)
        self.assertEqual(project_policies.filter(is_active=True).count(), 1)
        if connection.features.has_select_for_update:
            project_lock_queries = [
                sql
                for sql in executed_queries
                if "ORGS_PROJECT" in sql and "FOR UPDATE" in sql
            ]
            self.assertTrue(project_lock_queries)
